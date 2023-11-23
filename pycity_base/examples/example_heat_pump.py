#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Example of the heat pump (HP) class.
"""

from __future__ import division

import os
import numpy as np

import pycity_base.classes.supply.heat_pump as hp

import pycity_base.classes.timer
import pycity_base.classes.weather
import pycity_base.classes.prices
import pycity_base.classes.environment

import openpyxl


def run_example():
    # Create environment
    timer = pycity_base.classes.timer.Timer()
    weather = pycity_base.classes.weather.Weather(timer, use_TRY=True)
    prices = pycity_base.classes.prices.Prices()
    environment = pycity_base.classes.environment.Environment(timer, weather, prices)

    #  Heatpump data path
    src_path = os.path.dirname(os.path.dirname(__file__))
    hp_data_path = os.path.join(src_path, 'inputs', 'heat_pumps.xlsx')
    heatpumpData = openpyxl.load_workbook(hp_data_path, data_only=True)
    dimplex_LA12TU = heatpumpData["Dimplex_LA12TU"]

    # Size of the worksheet
    number_rows = dimplex_LA12TU.max_row
    number_columns = dimplex_LA12TU.max_column
    # Flow, ambient and max. temperatures
    t_flow = np.zeros(number_columns-2)
    t_ambient = np.zeros(int((number_rows-7)/2))
    t_max = dimplex_LA12TU.cell(1, 2).value

    firstRowCOP = number_rows - len(t_ambient)

    q_nominal = np.empty((len(t_ambient), len(t_flow)))
    cop = np.empty((len(t_ambient), len(t_flow)))

    for i in range(number_columns-2):
        t_flow[i] = dimplex_LA12TU.cell(4, 3+i).value

    for col in range(len(t_flow)):
        for row in range(len(t_ambient)):
            q_nominal[row, col] = dimplex_LA12TU.cell(int(5+row), int(3+col)).value
            cop[row, col] = dimplex_LA12TU.cell(int(firstRowCOP+row+1), int(3+col)).value

    p_nominal = q_nominal / cop

    # Create HP
    lower_activation_limit = 0.5

    heater = hp.Heatpump(environment, t_ambient, t_flow, q_nominal, p_nominal, cop, t_max, lower_activation_limit)

    # Print results
    print()
    print(("Type: " + heater.kind))
    print()
    print(("Maximum flow temperature: " + str(heater.t_max)))
    print(("Lower activation limit: " + str(heater.lower_activation_limit)))

    np.random.seed(0)
    flow_temperature = np.random.rand(timer.timesteps_horizon) * 20 + 35

    nominals = heater.getNominalValues(flow_temperature)
    print()
    print(("Nominal electricity consumption: " + str(nominals[0])))
    print(("Nominal heat output: " + str(nominals[1])))
    print(("Maximum flow temperature: " + str(nominals[2])))
    print(("Lower activation limit: " + str(nominals[3])))

    schedule = np.random.randint(2, size=timer.timesteps_used_horizon)
    result_p = (nominals[0])[0:96] * schedule
    result_q = (nominals[1])[0:96] * schedule

    heater.setResults(result_p, result_q, schedule)

    results = heater.getResults(True)
    print()
    print("Electricity input: " + str(results[0]))
    print()
    print("Heat output: " + str(results[1]))
    print()
    print("Schedule: " + str(results[2]))


if __name__ == '__main__':
    #  Run program
    run_example()
