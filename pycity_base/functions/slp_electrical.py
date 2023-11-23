#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 21 16:50:36 2015

@author: tsz
"""

from __future__ import division

import os
import numpy as np
import openpyxl
from pycity_base.functions import change_resolution as chres


def load(filename):
    # Open the workbook and get the sheet with all profiles
    book = openpyxl.load_workbook(filename, data_only=True)
    sheet = book["Profiles"]

    # Fill all values into one dictionary
    profiles = {}
    for c in range(1, sheet.max_column):
        # Get each key
        key = sheet.cell(1, c+1).value
        # Get the corresponding values
        values = [sheet.cell(r+1, c+1).value for r in range(2, sheet.max_row)]

        # Store the results
        profiles[key] = np.array(values)

    # Return the results
    return profiles


def get_demand(annual_demand, profile, time_discretization):
    scaling = 4000 / 1000000 * annual_demand / time_discretization * 900

    # Adjust sampling time
    if time_discretization != 900:
        changeResol = chres.changeResolution
        profile = changeResol(profile, 900, time_discretization, "sum")
        # Use scaling method "sum", because the values are given
        # in kWh for each 15 minutes

    return scaling * profile


if __name__ == "__main__":
    time_dis = 3600

    #  Define src path
    src_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Temperature - mean values
    filename = 'slp_electrical_2019.xlsx'
    import_path = os.path.join(src_path, 'inputs', 'standard_load_profile', filename)

    profiles = load(import_path, time_dis)

    load_household = get_demand(3000, profiles["H0"], time_dis)
